# programme.py
import logging
from collections import defaultdict

import bleach
from flask import Blueprint, render_template, redirect, url_for, request, flash, current_app, jsonify

import os

import html

import json
import re
import time
from flask_login import login_required, current_user

from ..forms import (
    CompetenceForm,
    DeleteForm,
    ReviewImportConfirmForm
)
# Import SQLAlchemy models
from ..models import (
    db,
    User,
    Programme,
    Competence,
    ElementCompetence,
    competence_programme,
    FilConducteur,
    Cours,
    CoursPrealable,
    CoursCorequis,
    ElementCompetence,
    ElementCompetenceCriteria,
    ElementCompetenceParCours,
    PlanDeCours
)
from flask import send_file
import io
from ...utils import generate_programme_grille_pdf

try:
    import openpyxl
    from openpyxl.styles import Alignment
except Exception:
    openpyxl = None
from ...utils.decorator import role_required, ensure_profile_completed, roles_required
from openai import OpenAI

# Utilities
# Example of another blueprint import (unused here, just as in your snippet)
logger = logging.getLogger(__name__)

programme_bp = Blueprint('programme', __name__, url_prefix='/programme')

@programme_bp.route('/<int:programme_id>/competences/logigramme')
@login_required
@ensure_profile_completed
def competence_logigramme(programme_id):
    """
    Vue interactive: logigramme des compétences reliées aux cours d'un programme,
    visualisant la progression par session et les liens développée/atteinte.
    """
    logger.debug(f"Accessing competence logigramme for programme ID: {programme_id}")
    programme = Programme.query.get_or_404(programme_id)

    # Vérifier l’accès
    if programme not in current_user.programmes and current_user.role != 'admin':
        flash("Vous n'avez pas accès à ce programme.", 'danger')
        return redirect(url_for('main.index'))

    # Collecte des cours + sessions via l'objet d'association CoursProgramme
    course_items = []
    course_ids = []
    for assoc in programme.cours_assocs:
        c = assoc.cours
        if not c:
            continue
        course_ids.append(c.id)
        course_items.append({
            'id': c.id,
            'code': c.code,
            'nom': c.nom,
            'session': assoc.session or 0,
            'fil_color': (c.fil_conducteur.couleur if getattr(c, 'fil_conducteur', None) and c.fil_conducteur.couleur else None),
            'fil_id': (c.fil_conducteur.id if getattr(c, 'fil_conducteur', None) else None),
            'fil_desc': (c.fil_conducteur.description if getattr(c, 'fil_conducteur', None) else None),
        })

    # Compétences associées au programme
    competences = (
        programme
        .competences
        .order_by(Competence.code)
        .all()
    )
    comp_items = [{'id': comp.id, 'code': comp.code, 'nom': comp.nom} for comp in competences]
    comp_ids = {c['id'] for c in comp_items}

    # Liens Compétence ↔ Cours
    # 1) Basé sur les éléments de compétence par cours et leur statut
    #    Statuts possibles: "Réinvesti", "Atteint", "Développé significativement"
    #    On agrège par (cours_id, competence_id) et on retient le statut dominant
    #    selon la priorité: Développé significativement > Atteint > Réinvesti.
    links = []
    if course_ids:
        from ..models import CompetenceParCours, ElementCompetenceParCours, ElementCompetence  # import local pour éviter cycles

        # Agrégation par (cours_id, competence_id)
        agg = {}
        q = (
            db.session.query(ElementCompetenceParCours.cours_id,
                             ElementCompetence.competence_id,
                             ElementCompetenceParCours.status)
            .join(ElementCompetence, ElementCompetenceParCours.element_competence_id == ElementCompetence.id)
            .filter(ElementCompetenceParCours.cours_id.in_(course_ids))
        )
        for cours_id, competence_id, status in q.all():
            if competence_id not in comp_ids:
                continue
            key = (cours_id, competence_id)
            d = agg.setdefault(key, {"reinvesti": 0, "atteint": 0, "developpe": 0, "total": 0})
            s = (status or '').strip().lower()
            if 'significativement' in s or 'développ' in s or 'developpe' in s:
                d['developpe'] += 1
            elif 'atteint' in s:
                d['atteint'] += 1
            elif 'réinvesti' in s or 'reinvesti' in s:
                d['reinvesti'] += 1
            d['total'] += 1

        # Construire les liens à partir de l'agrégation
        for (cours_id, competence_id), counts in agg.items():
            if counts['developpe'] > 0:
                ltype = 'developpe'
            elif counts['atteint'] > 0:
                ltype = 'atteint'
            elif counts['reinvesti'] > 0:
                ltype = 'reinvesti'
            else:
                continue
            links.append({
                'competence_id': competence_id,
                'cours_id': cours_id,
                'type': ltype,
                'weight': counts['total'],
                'counts': counts,
            })

        # 2) Compléter avec CompetenceParCours si aucun élément détaillé (fallback)
        rows = CompetenceParCours.query.filter(CompetenceParCours.cours_id.in_(course_ids)).all()
        seen = {(l['cours_id'], l['competence_id']) for l in links}
        for r in rows:
            if r.competence_developpee_id and r.competence_developpee_id in comp_ids:
                key = (r.cours_id, r.competence_developpee_id)
                if key not in seen:
                    links.append({
                        'competence_id': r.competence_developpee_id,
                        'cours_id': r.cours_id,
                        'type': 'developpe',
                        'weight': 1,
                        'counts': {'developpe': 1, 'atteint': 0, 'reinvesti': 0, 'total': 1}
                    })
            if r.competence_atteinte_id and r.competence_atteinte_id in comp_ids:
                key = (r.cours_id, r.competence_atteinte_id)
                if key not in seen:
                    links.append({
                        'competence_id': r.competence_atteinte_id,
                        'cours_id': r.cours_id,
                        'type': 'atteint',
                        'weight': 1,
                        'counts': {'developpe': 0, 'atteint': 1, 'reinvesti': 0, 'total': 1}
                    })

    # Sessions présentes (pour colonnes)
    sessions = sorted({int(c['session']) for c in course_items if c['session'] is not None})
    if not sessions:
        sessions = [0]

    # Dictionnaire des fils conducteurs présents
    fils = {}
    for c in course_items:
        if c.get('fil_id'):
            fid = c['fil_id']
            if fid not in fils:
                fils[fid] = {'id': fid, 'description': c.get('fil_desc'), 'couleur': c.get('fil_color')}

    data = {
        'programme': {'id': programme.id, 'nom': programme.nom},
        'competences': comp_items,
        'cours': course_items,
        'links': links,
        'sessions': sessions,
        'fils': list(fils.values())
    }

    # Génération IA: formulaire des modèles dispo
    try:
        from ..forms import GenerateContentForm
        generate_form = GenerateContentForm()
    except Exception:
        generate_form = None

    return render_template('programme/competence_logigramme.html', programme=programme, data=data, generate_form=generate_form)


@programme_bp.route('/<int:programme_id>/competences/logigramme/export.xlsx')
@login_required
@ensure_profile_completed
def export_competence_logigramme_xlsx(programme_id):
    """Export cours×compétences (statut) en XLSX.
    Lignes: cours (triés par session puis code). Colonnes: compétences (par code).
    Cellules: symbole selon statut agrégé pour (cours, compétence).
    """
    if openpyxl is None:
        return jsonify({'error': "Dépendance 'openpyxl' manquante côté serveur."}), 500
    programme = Programme.query.get_or_404(programme_id)
    if programme not in current_user.programmes and current_user.role != 'admin':
        flash("Vous n'avez pas accès à ce programme.", 'danger')
        return redirect(url_for('main.index'))

    # Collect cours (with session ordering)
    cours_list = []
    for assoc in programme.cours_assocs:
        if assoc.cours:
            cours_list.append((int(assoc.session or 0), assoc.cours.code or '', assoc.cours))
    cours_list.sort(key=lambda t: (t[0], t[1]))

    # Competences
    competences = programme.competences.order_by(Competence.code).all()
    comp_ids = {c.id for c in competences}

    # Aggregate statuses for each (cours, competence)
    from ..models import CompetenceParCours, ElementCompetenceParCours, ElementCompetence
    course_ids = [c.id for _,__, c in cours_list]
    agg = {}
    if course_ids:
        q = (
            db.session.query(ElementCompetenceParCours.cours_id,
                             ElementCompetence.competence_id,
                             ElementCompetenceParCours.status)
            .join(ElementCompetence, ElementCompetenceParCours.element_competence_id == ElementCompetence.id)
            .filter(ElementCompetenceParCours.cours_id.in_(course_ids))
        )
        for cours_id, competence_id, status in q.all():
            if competence_id not in comp_ids:
                continue
            key = (cours_id, competence_id)
            d = agg.setdefault(key, {"reinvesti": 0, "atteint": 0, "developpe": 0, "total": 0})
            s = (status or '').strip().lower()
            if 'significativement' in s or 'développ' in s or 'developpe' in s:
                d['developpe'] += 1
            elif 'atteint' in s:
                d['atteint'] += 1
            elif 'réinvesti' in s or 'reinvesti' in s:
                d['reinvesti'] += 1
            d['total'] += 1
        # Fallback CompetenceParCours
        rows = CompetenceParCours.query.filter(CompetenceParCours.cours_id.in_(course_ids)).all()
        seen = set(agg.keys())
        for r in rows:
            if r.competence_developpee_id and r.competence_developpee_id in comp_ids and (r.cours_id, r.competence_developpee_id) not in seen:
                agg[(r.cours_id, r.competence_developpee_id)] = {"developpe": 1, "atteint": 0, "reinvesti": 0, "total": 1}
            if r.competence_atteinte_id and r.competence_atteinte_id in comp_ids and (r.cours_id, r.competence_atteinte_id) not in seen:
                agg[(r.cours_id, r.competence_atteinte_id)] = {"developpe": 0, "atteint": 1, "reinvesti": 0, "total": 1}

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Logigramme'
    # Header row
    ws.cell(row=1, column=1, value='Cours\\Compétences')
    for j, comp in enumerate(competences, start=2):
        ws.cell(row=1, column=j, value=comp.code)
    # Symbols map
    sym_map = { 'developpe': '•', 'atteint': '✓', 'reinvesti': '↺' }
    # Fill rows
    for i, (_, __, cours) in enumerate(cours_list, start=2):
        ws.cell(row=i, column=1, value=cours.code)
        for j, comp in enumerate(competences, start=2):
            d = agg.get((cours.id, comp.id))
            if not d:
                continue
            if d['developpe'] > 0:
                t = 'developpe'
            elif d['atteint'] > 0:
                t = 'atteint'
            else:
                t = 'reinvesti'
            ws.cell(row=i, column=j, value=sym_map.get(t, ''))
    # Align center
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    # Autosize columns (rough)
    for col in ws.columns:
        maxlen = 10
        for cell in col:
            try:
                v = str(cell.value) if cell.value is not None else ''
                if len(v) > maxlen:
                    maxlen = len(v)
            except Exception:
                pass
        ws.column_dimensions[col[0].column_letter].width = maxlen + 2

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    filename = f"logigramme_{programme_id}.xlsx"
    return send_file(bio, as_attachment=True, download_name=filename, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@programme_bp.route('/<int:programme_id>/links', methods=['POST'])
@login_required
@ensure_profile_completed
def save_programme_links(programme_id):
    """Persist edited links into ElementCompetenceParCours rows.
    For each provided link {cours_id, competence_id, type, weight}, we set the
    status of ElementCompetenceParCours rows for that (cours, competence)'s elements
    to the selected type. If no rows exist, we create rows for the competence's
    elements for this course. Weight is not strictly enforced at DB level because
    the number of elements is authoritative; all rows are set to the given type.
    """
    from flask import jsonify
    programme = Programme.query.get_or_404(programme_id)
    # Restrict to admin/coordo only
    if current_user.role not in ('admin', 'coordo'):
        return jsonify({'error': "Accès restreint (admin/coordo)"}), 403
    # If not admin, ensure membership to programme
    if current_user.role != 'admin' and programme not in current_user.programmes:
        return jsonify({'error': "Accès refusé"}), 403

    try:
        payload = (request.get_json(silent=True) or [])
        assert isinstance(payload, list)
    except Exception:
        return jsonify({'error': 'Format JSON invalide'}), 400

    # Build quick lookups for validation
    course_ids = {assoc.cours_id for assoc in programme.cours_assocs}
    comp_ids = {c.id for c in programme.competences}
    allowed_types = {'developpe': 'Développé significativement', 'atteint': 'Atteint', 'reinvesti': 'Réinvesti'}

    # Build competence elements index for this programme
    comp_ids = {c.id for c in programme.competences}
    # Map competence_id -> [element ids]
    comp_to_elem_ids = {}
    all_elem_ids = []
    for cid in comp_ids:
      ids = [e.id for e in ElementCompetence.query.filter_by(competence_id=cid).all()]
      comp_to_elem_ids[cid] = ids
      all_elem_ids.extend(ids)

    # Build payload pairs set
    payload_pairs = set()
    updated = 0
    for item in payload:
        try:
            cours_id = int(item.get('cours_id'))
            competence_id = int(item.get('competence_id'))
            ltype = str(item.get('type'))
        except Exception:
            continue
        if cours_id not in course_ids or competence_id not in comp_ids:
            continue
        if ltype not in allowed_types:
            continue
        payload_pairs.add((cours_id, competence_id))
        status_value = allowed_types[ltype]

        # Fetch elements for this competence
        elements = ElementCompetence.query.filter_by(competence_id=competence_id).order_by(ElementCompetence.id).all()
        if not elements:
            # No elements defined; skip silently
            continue

        # Ensure one EPC row per element for this course
        existing = { (epc.element_competence_id): epc for epc in ElementCompetenceParCours.query
                     .filter_by(cours_id=cours_id)
                     .filter(ElementCompetenceParCours.element_competence_id.in_([e.id for e in elements]))
                     .all() }
        for e in elements:
            epc = existing.get(e.id)
            if not epc:
                epc = ElementCompetenceParCours(cours_id=cours_id, element_competence_id=e.id, status=status_value)
                db.session.add(epc)
            else:
                epc.status = status_value
        updated += 1

    # Handle deletions: remove EPC rows for pairs not present in payload
    if all_elem_ids:
        # Fetch all EPC rows for programme courses and competence elements
        rows = (ElementCompetenceParCours.query
                .filter(ElementCompetenceParCours.cours_id.in_(course_ids))
                .filter(ElementCompetenceParCours.element_competence_id.in_(all_elem_ids))
                .all())
        # Build element->competence map
        elem_to_comp = {}
        for cid, eids in comp_to_elem_ids.items():
            for eid in eids:
                elem_to_comp[eid] = cid
        for r in rows:
            comp_id = elem_to_comp.get(r.element_competence_id)
            if comp_id is None:
                continue
            if (r.cours_id, comp_id) not in payload_pairs:
                db.session.delete(r)

    try:
        db.session.commit()
    except Exception as ex:
        db.session.rollback()
        current_app.logger.exception("Error saving links")
        return jsonify({'error': "Erreur lors de l'enregistrement"}), 500

    return jsonify({'ok': True, 'updated': updated, 'message': f'{updated} liens appliqués'}), 200


@programme_bp.route('/<int:programme_id>/logigramme/generate', methods=['POST'])
@login_required
@ensure_profile_completed
def generate_competence_logigramme(programme_id):
    """Start Celery task to generate suggested links; returns task_id for polling."""
    programme = Programme.query.get_or_404(programme_id)
    if current_user.role not in ('admin', 'coordo'):
        return jsonify({'error': 'Accès restreint (admin/coordo)'}), 403
    if current_user.role != 'admin' and programme not in current_user.programmes:
        return jsonify({'error': 'Accès refusé'}), 403
    user = current_user
    if not user.openai_key:
        return jsonify({'error': 'Aucune clé OpenAI configurée dans votre profil.'}), 400
    if not user.credits or user.credits <= 0:
        return jsonify({'error': "Crédits insuffisants pour effectuer l'appel."}), 400
    form = request.get_json(silent=True) or {}

    # Enqueue task
    from ..tasks.generation_logigramme import generate_programme_logigramme_task
    task = generate_programme_logigramme_task.delay(programme_id, user.id, form)
    return jsonify({'task_id': task.id}), 202


    


@programme_bp.route('/api/task_status/<task_id>')
@login_required
def programme_task_status(task_id):
    from ...celery_app import celery
    from celery.result import AsyncResult
    try:
        task_result = AsyncResult(task_id, app=celery)
        state = task_result.state
        result = task_result.result
        resp = { 'task_id': task_id, 'state': state }
        if state == 'PROGRESS':
            info = task_result.info or {}
            if isinstance(info, dict):
                resp.update(info)
        elif state == 'SUCCESS':
            if isinstance(result, dict):
                resp.update(result)
            else:
                resp.update({ 'status': 'success', 'result': result })
        elif state == 'FAILURE':
            resp.update({ 'status': 'error', 'message': str(result) if result else 'Echec de la tâche' })
        return jsonify(resp)
    except Exception as e:
        current_app.logger.exception('Erreur statut de tâche')
        return jsonify({ 'task_id': task_id, 'state': 'ERROR', 'status': 'error', 'message': str(e) }), 500

@programme_bp.route('/<int:programme_id>/competences')
@login_required
@ensure_profile_completed
def view_competences_programme(programme_id):
    """
    Affiche la liste de toutes les compétences associées à un programme spécifique.
    """
    logger.debug(f"Accessing competencies list for programme ID: {programme_id}")
    programme = Programme.query.get_or_404(programme_id)

    # Vérifier l’accès
    if programme not in current_user.programmes and current_user.role != 'admin':
        flash("Vous n'avez pas accès à ce programme.", 'danger')
        return redirect(url_for('main.index'))

    # Récupérer toutes les compétences via la relation many-to-many,
    # triées par code
    competences = (
        programme
        .competences          # backref dynamique défini sur Competence.programmes
        .order_by(Competence.code)
        .all()
    )

    return render_template(
        'programme/view_competences_programme.html',
        programme=programme,
        competences=competences
    )


@programme_bp.route('/review_import', methods=['POST'])
@login_required
def review_competencies_import():
    """
    Prépare un comparatif côte-à-côte pour chaque compétence identifiée par son code,
    entre la version en base de données et celle extraite du fichier JSON, en
    particulier pour le champ "Contexte" qui sera transformé pour ressembler à la version BD.
    """
    # Récupération des paramètres
    programme_id = request.form.get('programme_id')
    base_filename = request.form.get('base_filename')
    if not programme_id or not base_filename:
        flash("Informations manquantes (programme ou devis) pour démarrer la révision.", "danger")
        return redirect(url_for('ocr.show_trigger_page'))
    
    try:
        programme_id = int(programme_id)
    except ValueError:
        flash("ID de programme invalide.", "danger")
        return redirect(url_for('ocr.show_trigger_page'))
    
    programme = db.session.get(Programme, programme_id)
    if not programme:
        flash(f"Programme avec l'ID {programme_id} non trouvé.", "danger")
        abort(404)
    
    # Chemins vers les fichiers
    txt_output_dir = current_app.config.get('TXT_OUTPUT_DIR')
    ocr_file_path = os.path.join(txt_output_dir, f"{base_filename}_ocr.md")
    competencies_file_path = os.path.join(txt_output_dir, f"{base_filename}_competences.json")
    
    # Lecture du fichier OCR (bien que non utilisé dans le comparatif)
    ocr_text = ""
    try:
        with open(ocr_file_path, 'r', encoding='utf-8') as f:
            ocr_text = f.read()
    except Exception as e:
        logger.error(f"Erreur lors de la lecture OCR {ocr_file_path}: {e}", exc_info=True)
    
    # Lecture du fichier JSON de compétences (avec stratégie de repli si le fichier exact n'existe pas)
    competencies_data = None
    has_competencies_file = False
    try:
        with open(competencies_file_path, 'r', encoding='utf-8') as f:
            competencies_data = json.load(f)
        if not isinstance(competencies_data, dict) or 'competences' not in competencies_data:
            raise ValueError("Structure JSON invalide: clé 'competences' manquante.")
        has_competencies_file = True
    except Exception as e:
        logger.warning(f"Lecture initiale échouée pour {competencies_file_path}: {e}")
        # Tentative de repli: rechercher un fichier _competences.json correspondant dans le dossier
        try:
            candidates = []
            if os.path.isdir(txt_output_dir):
                base_lower = base_filename.lower()
                for fname in os.listdir(txt_output_dir):
                    if not fname.endswith('_competences.json'):
                        continue
                    f_lower = fname.lower()
                    # Correspondances permissives: préfixe identique ou inclusion du code de programme
                    if f_lower.startswith(base_lower) or base_lower.startswith(f_lower.replace('_competences.json','')):
                        candidates.append(os.path.join(txt_output_dir, fname))
                # Si aucune correspondance stricte, tenter une correspondance contenant le code (avant le premier '_')
                if not candidates and '_' in base_lower:
                    code_prefix = base_lower.split('_', 1)[0]
                    for fname in os.listdir(txt_output_dir):
                        if fname.lower().endswith('_competences.json') and fname.lower().startswith(code_prefix):
                            candidates.append(os.path.join(txt_output_dir, fname))
            if len(candidates) == 1:
                fallback_path = candidates[0]
                logger.info(f"Utilisation du JSON de repli détecté: {fallback_path}")
                competencies_file_path = fallback_path
                with open(competencies_file_path, 'r', encoding='utf-8') as f:
                    competencies_data = json.load(f)
                if not isinstance(competencies_data, dict) or 'competences' not in competencies_data:
                    raise ValueError("Structure JSON invalide dans le fichier de repli: clé 'competences' manquante.")
                has_competencies_file = True
                flash("Fichier des compétences introuvable sous le nom attendu; utilisation d'une correspondance trouvée.", "info")
            elif len(candidates) > 1:
                logger.warning(f"Plusieurs fichiers candidats trouvés pour base '{base_filename}': {candidates}")
                flash("Plusieurs fichiers de compétences possibles trouvés. Merci de relancer en sélectionnant le bon devis.", "warning")
            else:
                logger.error(f"Aucun fichier JSON correspondant trouvé dans {txt_output_dir} pour base '{base_filename}'.")
                flash(f"Erreur lors de la lecture ou du parsing du fichier JSON des compétences: {e}", "warning")
        except Exception as e2:
            logger.error(f"Erreur lors de la recherche/sélection d'un fichier JSON de repli: {e2}", exc_info=True)
            flash(f"Erreur lors de la lecture ou du parsing du fichier JSON des compétences: {e}", "warning")
    
    # Construction du comparatif pour chaque compétence
    comparisons = []
    if has_competencies_file and competencies_data and "competences" in competencies_data:
        for comp in competencies_data["competences"]:
            # Récupérer le code (dans "code" ou "Code")
            code = comp.get("code") or comp.get("Code")
            if not code:
                continue
            # Recherche en BD par code et programme
            db_comp = Competence.query.filter_by(code=code, programme_id=programme.id).first()
            db_version = None
            if db_comp:
                db_elements = []
                for elem in db_comp.elements:
                    criteria_list = [crit.criteria for crit in elem.criteria] if elem.criteria else []
                    db_elements.append({
                        "nom": elem.nom,
                        "criteres": criteria_list
                    })
                # Conversion du champ texte en une liste de critères à partir des lignes du texte
                db_criteres = db_comp.criteria_de_performance

                db_version = {
                    "code": db_comp.code,
                    "nom": db_comp.nom,
                    "contexte": db_comp.contexte_de_realisation,
                    "criteres": db_criteres,
                    "elements": db_elements
                }
            # Traitement de la version JSON
            json_elements = []
            if comp.get("Éléments"):
                for elem in comp["Éléments"]:
                    if isinstance(elem, str):
                        json_elements.append({
                            "nom": elem,
                            "criteres": None
                        })
                    elif isinstance(elem, dict):
                        json_elements.append({
                            "nom": elem.get("element") or elem.get("nom"),
                            "criteres": elem.get("criteres")
                        })
            json_version = {
                "code": comp.get("Code") or comp.get("code"),
                "nom": comp.get("Nom de la compétence") or comp.get("nom"),
                "contexte": comp.get("Contexte de réalisation"),
                "criteres": comp.get("Critères de performance pour l’ensemble de la compétence"),
                "elements": json_elements
            }
            comparisons.append({
                "code": code,
                "db": db_version,
                "json": json_version
            })
    
    form = ReviewImportConfirmForm()
    form.programme_id.data = programme_id
    form.base_filename.data = base_filename
    form.import_structured.data = 'true' if has_competencies_file and competencies_data else 'false'
    
    return render_template('programme/review_import.html',
                           programme=programme,
                           comparisons=comparisons,
                           base_filename=base_filename,
                           form=form)


def json_to_html_list(data):
    """
    Convertit une structure JSON (liste ou dict avec 'texte'/'sous_points')
    en une chaîne HTML formatée en liste imbriquée, en échappant le contenu texte.
    """
    if isinstance(data, list):
        # If data is a list, process each item and wrap in <ul>
        if not data: # Handle empty list
            return ""
        html_items = ""
        for item in data:
            html_items += json_to_html_list(item) # Recursive call for each item
        return f"<ul>{html_items}</ul>"

    elif isinstance(data, dict):
        # If data is a dict, extract 'texte', escape it, start <li>,
        # recursively process 'sous_points' if they exist, then close </li>
        texte = data.get("texte") or data.get("Text") or ""
        escaped_texte = html.escape(texte) # <-- Escape the text content
        html_content = f"<li>{escaped_texte}" # Start list item with escaped text

        sous_points = data.get("sous_points")
        if isinstance(sous_points, list) and sous_points: # Check if sous_points is a non-empty list
            # Recursive call for sub-points - this will return a nested <ul>...</ul>
            html_content += json_to_html_list(sous_points)

        html_content += "</li>" # Close the list item
        return html_content

    elif isinstance(data, str):
        # If data is just a string, escape it and wrap it in <li>
        if not data.strip(): # Handle empty string
             return ""
        escaped_data = html.escape(data) # <-- Escape the string content
        return f"<li>{escaped_data}</li>"

    else:
        # Ignore other data types or return their string representation (escaped)
        return html.escape(str(data)) if data is not None else ""

def format_context(context):
    """
    Formats the contexte de réalisation for database storage, aiming for safe HTML.

    - If 'context' is a list (typically from JSON like [{'texte': ..., 'sous_points': ...}]),
      it delegates to json_to_html_list to create potentially nested HTML <ul>/<li> lists.
    - If 'context' is a dictionary, it attempts to extract the 'texte' key or provides
      a safe string representation.
    - If 'context' is a string containing newline characters, it formats it as an
      HTML <ul> list with each line as an <li> item.
    - If 'context' is a simple string, it returns the HTML-escaped string.
    - Handles other types by converting them to string and escaping.

    Args:
        context: The data to format (list, dict, str, etc.).

    Returns:
        A string containing safe HTML representation of the context.
    """
    if isinstance(context, list):
        # Delegate list processing to the recursive function that handles nesting
        # Assumes json_to_html_list exists and handles escaping internally
        try:
            # Add try-except block for robustness if json_to_html_list might fail
             return json_to_html_list(context)
        except Exception as e:
             # Log the error appropriately in a real application
             print(f"Error processing list context with json_to_html_list: {e}")
             # Fallback to a safe representation
             return html.escape(str(context))


    elif isinstance(context, dict):
        # For a dictionary, try to get 'texte' key, otherwise use string representation. Escape result.
        # This is a fallback; ideally, dicts should be part of lists handled above.
        text_content = context.get('texte', str(context))
        return html.escape(text_content)

    elif isinstance(context, str):
        # If the string contains newline characters, treat each line as a list item
        if "\n" in context:
            lignes = [ligne.strip() for ligne in context.splitlines() if ligne.strip()]
            # Escape each line before wrapping in <li>
            escaped_lignes = [f"<li>{html.escape(ligne)}</li>" for ligne in lignes]
            # Return as a <ul> list if there are any lines
            return f"<ul>{''.join(escaped_lignes)}</ul>" if escaped_lignes else ""
        else:
            # For a simple string (no newlines), just strip whitespace and escape it
            return html.escape(context.strip())

    # Handle other potential types (int, float, None, etc.) safely
    elif context is None:
        return "" # Return empty string for None
    else:
        # Convert any other type to string and escape it
        return html.escape(str(context))

def format_global_criteria(text: str) -> str:
    """
    Convertit une chaîne de critères (une ligne par critère) en une liste HTML.
    Par exemple:
        "Critère 1.\nCritère 2."
    devient:
        <ul>
            <li>Critère 1.</li>
            <li>Critère 2.</li>
        </ul>
    """
    # Supprimer les espaces et ignorer les lignes vides
    criteria_lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not criteria_lines:
        return ""
    # Construit la liste HTML
    html_list = "<ul>\n" + "\n".join(f"\t<li>{line}</li>" for line in criteria_lines) + "\n</ul>"
    return html_list

@programme_bp.route('/confirm_import', methods=['POST'])
@login_required
def confirm_competencies_import():
    """
    Traite la confirmation depuis la page de révision et importe les compétences en base de données.
    Conversion du champ "contexte" du JSON en HTML dans le format souhaité.
    """
    form = ReviewImportConfirmForm()
    if form.validate_on_submit():
        programme_id = form.programme_id.data
        base_filename = form.base_filename.data
        import_structured = form.import_structured.data == 'true'

        # Récupération du programme cible
        try:
            programme = db.session.get(Programme, int(programme_id))
        except ValueError:
            flash("ID de programme invalide.", "danger")
            return redirect(url_for('main.index'))

        if not programme:
            flash("Programme cible non trouvé lors de la confirmation.", "danger")
            return redirect(url_for('main.index'))

        # Cas 1: Seule confirmation du texte OCR
        if not import_structured:
            flash(
                f"Texte OCR pour '{base_filename}' confirmé et associé au programme '{programme.nom}'. Aucune compétence importée.",
                "info"
            )
            return redirect(url_for('programme.view_programme', programme_id=programme.id))

        # Cas 2: Importation des compétences à partir du fichier JSON
        txt_output_dir = current_app.config.get('TXT_OUTPUT_DIR')
        competencies_file_path = os.path.join(txt_output_dir, f"{base_filename}_competences.json")
        competences_added_count = 0
        elements_added_count = 0
        competences_updated_count = 0

        try:
            logger.info(f"Début de l'importation depuis {competencies_file_path} pour le Programme ID {programme.id}")
            with open(competencies_file_path, 'r', encoding='utf-8') as f:
                competencies_data = json.load(f)

            competences_list = competencies_data.get('competences', [])
            if not isinstance(competences_list, list):
                raise ValueError("La clé 'competences' dans le JSON n'est pas une liste.")

            for comp_data in competences_list:
                if not isinstance(comp_data, dict):
                    logger.warning(f"Entrée compétence invalide (pas un dict): {comp_data}")
                    continue

                # Extraction des données en essayant plusieurs clés selon le JSON
                code = comp_data.get('code') or comp_data.get('Code')
                nom = comp_data.get('enonce') or comp_data.get('nom') or comp_data.get('Nom de la compétence')

                # Extraction et formatage du contexte
                raw_contexte = (comp_data.get('contexte_de_realisation') or 
                                comp_data.get('Contexte') or 
                                comp_data.get('Contexte de réalisation'))
                contexte = ""
                if raw_contexte is not None:
                    contexte = format_context(raw_contexte)

                # Extraction des critères de performance globaux
                critere_perf = (comp_data.get('criteria_de_performance') or 
                                comp_data.get('Critères de performance pour l’ensemble de la compétence') or 
                                comp_data.get('Critères de performance'))
                if critere_perf is not None:
                    # S'il s'agit d'une liste, on la convertit en chaîne avec un retour à la ligne
                    if isinstance(critere_perf, list):
                        critere_perf = "\n".join(critere_perf)
                    elif isinstance(critere_perf, str):
                        critere_perf = critere_perf.strip()
                    # On transforme la chaîne en liste HTML
                    critere_perf = format_global_criteria(critere_perf)

                if not code or not nom:
                    logger.warning(f"Compétence ignorée (code ou nom manquant): {comp_data}")
                    continue

                # Recherche d'une compétence existante dans la base
                existing_comp = Competence.query.filter_by(code=code, programme_id=programme.id).first()
                if existing_comp:
                    existing_comp.nom = nom
                    existing_comp.contexte_de_realisation = contexte
                    existing_comp.criteria_de_performance = critere_perf
                    db.session.add(existing_comp)
                    competences_updated_count += 1
                    logger.info(f"Compétence {code} mise à jour.")
                    comp_to_process = existing_comp
                else:
                    new_comp = Competence(
                        code=code,
                        nom=nom,
                        contexte_de_realisation=contexte,
                        criteria_de_performance=critere_perf,
                        programme_id=programme.id
                    )
                    db.session.add(new_comp)
                    db.session.flush()  # Pour obtenir l'ID si nécessaire pour les éléments
                    comp_to_process = new_comp
                    competences_added_count += 1
                    logger.info(f"Création de la compétence {code}.")

                # Associer la compétence au programme dans la table d'association (utilisée par le logigramme)
                try:
                    link_exists = db.session.execute(
                        db.text(
                            "SELECT 1 FROM Competence_Programme WHERE competence_id = :cid AND programme_id = :pid"
                        ),
                        {"cid": comp_to_process.id, "pid": programme.id}
                    ).first()
                    if not link_exists:
                        db.session.execute(
                            db.text(
                                "INSERT INTO Competence_Programme (competence_id, programme_id) VALUES (:cid, :pid)"
                            ),
                            {"cid": comp_to_process.id, "pid": programme.id}
                        )
                        logger.debug(f"Association Competence[{comp_to_process.id}] -> Programme[{programme.id}] créée.")
                except Exception as link_err:
                    logger.warning(f"Impossible de créer l'association Competence_Programme pour {code}: {link_err}")

                # Traitement complet des éléments de compétence :
                # Récupération de la liste d'éléments depuis le JSON
                elements_list = comp_data.get('elements') or comp_data.get('Éléments') or []
                # Construction d'un dictionnaire associant chaque nom d'élément à ses données JSON (incluant les critères)
                json_elements = {}
                for elem_data in elements_list:
                    if isinstance(elem_data, str):
                        nom_elem = elem_data.strip()
                        json_elements[nom_elem] = {"criteres": None}
                    elif isinstance(elem_data, dict):
                        nom_elem = elem_data.get('element') or elem_data.get('nom')
                        if not nom_elem:
                            continue
                        json_elements[nom_elem] = {
                            "criteres": elem_data.get('criteres')
                        }
                
                # Synchronisation des éléments en base
                current_elements = comp_to_process.elements if hasattr(comp_to_process, 'elements') else []
                # Suppression des éléments en base qui ne sont plus présents dans le JSON
                for elem in list(current_elements):
                    if elem.nom not in json_elements:
                        db.session.delete(elem)
                        logger.debug(f"Suppression de l'élément '{elem.nom}' pour la compétence {code}.")

                # Mise à jour ou création des éléments et de leurs critères
                current_elements_names = {elem.nom for elem in comp_to_process.elements} if comp_to_process.elements else set()
                for nom_elem, elem_info in json_elements.items():
                    json_criteres = elem_info.get("criteres")
                    if nom_elem in current_elements_names:
                        # Mise à jour des critères de l'élément existant
                        elem = next(e for e in comp_to_process.elements if e.nom == nom_elem)
                        elem.criteria.clear()  # Suppression des critères existants
                        if json_criteres:
                            if isinstance(json_criteres, list):
                                for crit in json_criteres:
                                    if isinstance(crit, (dict, list)):
                                        crit_text = json_to_html_list(crit)
                                    else:
                                        crit_text = str(crit).strip()
                                    new_crit = ElementCompetenceCriteria(criteria=crit_text)
                                    elem.criteria.append(new_crit)
                            else:
                                if isinstance(json_criteres, (dict, list)):
                                    crit_text = json_to_html_list(json_criteres)
                                else:
                                    crit_text = str(json_criteres).strip()
                                new_crit = ElementCompetenceCriteria(criteria=crit_text)
                                elem.criteria.append(new_crit)
                        logger.debug(f"Mise à jour des critères de l'élément '{nom_elem}' pour la compétence {code}.")
                    else:
                        # Création d'un nouvel élément avec ses critères
                        new_elem = ElementCompetence(nom=nom_elem, competence_id=comp_to_process.id)
                        if json_criteres:
                            if isinstance(json_criteres, list):
                                for crit in json_criteres:
                                    if isinstance(crit, (dict, list)):
                                        crit_text = json_to_html_list(crit)
                                    else:
                                        crit_text = str(crit).strip()
                                    new_crit = ElementCompetenceCriteria(criteria=crit_text)
                                    new_elem.criteria.append(new_crit)
                            else:
                                if isinstance(json_criteres, (dict, list)):
                                    crit_text = json_to_html_list(json_criteres)
                                else:
                                    crit_text = str(json_criteres).strip()
                                new_crit = ElementCompetenceCriteria(criteria=crit_text)
                                new_elem.criteria.append(new_crit)
                        db.session.add(new_elem)
                        elements_added_count += 1
                        logger.debug(f"Création de l'élément '{nom_elem}' avec ses critères pour la compétence {code}.")

            db.session.commit()
            flash_msg = f"Import terminé pour le programme '{programme.nom}'."
            if competences_added_count:
                flash_msg += f" {competences_added_count} compétence(s) ajoutée(s)."
            if competences_updated_count:
                flash_msg += f" {competences_updated_count} compétence(s) mise(s) à jour."
            if elements_added_count:
                flash_msg += f" {elements_added_count} élément(s) ajouté(s)/supprimé(s)."
            if not (competences_added_count or competences_updated_count or elements_added_count):
                flash_msg += " Aucune donnée nouvelle à importer."
            flash(flash_msg, "success")
        except FileNotFoundError:
            db.session.rollback()
            flash(f"Fichier JSON '{os.path.basename(competencies_file_path)}' non trouvé lors de l'import.", "danger")
        except Exception as e:
            db.session.rollback()
            flash(f"Erreur lors de l'importation en base de données: {e}", "danger")
            logger.error(f"Erreur inattendue lors de l'importation pour le programme {programme.id}: {e}", exc_info=True)

        return redirect(url_for('programme.view_programme', programme_id=programme.id))
    else:
        flash("Erreur lors de la soumission du formulaire de confirmation. Veuillez réessayer.", "danger")
        try:
            prog_id = int(request.form.get('programme_id', 0))
            return redirect(url_for('programme.view_programme', programme_id=prog_id))
        except Exception:
            return redirect(url_for('main.index'))



@programme_bp.route('/<int:programme_id>')
@login_required
@ensure_profile_completed
def view_programme(programme_id):
    # Debug logging
    logger.debug(f"Accessing programme {programme_id}")
    logger.debug(f"User programmes: {[p.id for p in current_user.programmes]}")
    
    # Récupérer le programme
    programme = db.session.get(Programme, programme_id)
    if not programme:
        flash('Programme non trouvé.', 'danger')
        return redirect(url_for('main.index'))

    # Vérifier si l'utilisateur a accès à ce programme
    if programme not in current_user.programmes and current_user.role != 'admin':
        flash("Vous n'avez pas accès à ce programme.", 'danger')
        return render_template('no_access.html')


    # Récupérer les compétences associées
    competences = Competence.query.filter_by(programme_id=programme_id).all()

    # Récupérer les fils conducteurs associés
    fil_conducteurs = FilConducteur.query.filter_by(programme_id=programme_id).all()

    # Récupérer les cours associés via la table d'association, triés par session
    # (chaque association CoursProgramme contient la session pour ce programme)
    # 'programme' a déjà été récupéré ci-dessus
    cours_assocs = sorted(programme.cours_assocs, key=lambda a: a.session)
    # Liste des objets Cours
    cours_liste = [assoc.cours for assoc in cours_assocs]
    # Mapping cours_id -> session spécifique à ce programme
    session_map = {assoc.cours.id: assoc.session for assoc in cours_assocs}


    # Regrouper les cours par session
    cours_par_session = defaultdict(list)
    
    # Pour chaque cours, récupérer les informations du dernier plan
    for cours in cours_liste:
        # Récupérer le dernier plan de cours
        dernier_plan = PlanDeCours.query\
            .filter_by(cours_id=cours.id)\
            .order_by(PlanDeCours.modified_at.desc())\
            .first()
        
        if dernier_plan:
            # Récupérer le username de l'utilisateur si modified_by_id existe
            modified_username = None
            if dernier_plan.modified_by_id:
                user = db.session.get(User, dernier_plan.modified_by_id)
                modified_username = user.username if user else None

            cours.dernier_plan = {
                'session': dernier_plan.session,
                'modified_at': dernier_plan.modified_at,
                'modified_by': modified_username
            }
        else:
            cours.dernier_plan = None
            
        # Grouper par session définie dans la relation CoursProgramme
        sess = session_map.get(cours.id)
        cours_par_session[sess].append(cours)

    # Récupérer préalables et co-requis pour chaque cours
    prerequisites = {}
    corequisites = {}
    for cours in cours_liste:  # Utiliser cours_liste au lieu de cours
        # Pré-requis
        preq = (db.session.query(Cours.nom, Cours.code, CoursPrealable.note_necessaire)
                .join(CoursPrealable, Cours.id == CoursPrealable.cours_prealable_id)
                .filter(CoursPrealable.cours_id == cours.id)
                .all())
        prerequisites[cours.id] = [(f"{p.code} - {p.nom}", p.note_necessaire) for p in preq]

        # Co-requis
        coreq = (db.session.query(Cours.nom, Cours.code)
                 .join(CoursCorequis, Cours.id == CoursCorequis.cours_corequis_id)
                 .filter(CoursCorequis.cours_id == cours.id)
                 .all())
        corequisites[cours.id] = [f"{cc.code} - {cc.nom}" for cc in coreq]

    # Récupérer les codes des compétences (développées ou atteintes) par cours
    competencies_codes = {}
    for c in cours_liste    :
        # Un SELECT DISTINCT sur c.code AS competence_code depuis la table Competence 
        # via ElementCompetence -> ElementCompetenceParCours
        comps = (db.session.query(Competence.code.label('competence_code'))
                 .join(ElementCompetence, Competence.id == ElementCompetence.competence_id)
                 .join(ElementCompetenceParCours, ElementCompetence.id == ElementCompetenceParCours.element_competence_id)
                 .filter(ElementCompetenceParCours.cours_id == c.id)
                 .filter(ElementCompetenceParCours.status.in_(['Développé significativement', 'Atteint']))
                 .distinct()
                 .all())
        competencies_codes[c.id] = [comp.competence_code for comp in comps]

    # Calcul des totaux
    total_heures_theorie = sum(c.heures_theorie for c in cours_liste)
    total_heures_laboratoire = sum(c.heures_laboratoire for c in cours_liste)
    total_heures_travail_maison = sum(c.heures_travail_maison for c in cours_liste)
    total_unites = sum(c.nombre_unites for c in cours_liste)

    # Créer des dictionnaires de formulaires de suppression
    delete_forms_competences = {comp.id: DeleteForm(prefix=f"competence-{comp.id}") for comp in competences}
    delete_forms_cours = {c.id: DeleteForm(prefix=f"cours-{c.id}") for c in cours_liste}

    # Récupérer tous les programmes (pour le sélecteur éventuel)
    programmes = current_user.programmes

    cours_plans_mapping = {}
    for cours in cours_liste:
        plans = PlanDeCours.query.filter_by(cours_id=cours.id).all()
        cours_plans_mapping[cours.id] = [p.to_dict() for p in plans]

    # Génération IA: liste de modèles (pour le modal)
    try:
        from ..forms import GenerateContentForm
        generate_form = GenerateContentForm()
    except Exception:
        generate_form = None

    return render_template('view_programme.html',
                           programme=programme,
                           programmes=programmes,
                           competences=competences,
                           fil_conducteurs=fil_conducteurs,
                           cours_par_session=cours_par_session,
                           delete_forms_competences=delete_forms_competences,
                           delete_forms_cours=delete_forms_cours,
                           prerequisites=prerequisites,
                           corequisites=corequisites,
                           competencies_codes=competencies_codes,
                           total_heures_theorie=total_heures_theorie,
                           total_heures_laboratoire=total_heures_laboratoire,
                           total_heures_travail_maison=total_heures_travail_maison,
                           total_unites=total_unites,
                           cours_plans_mapping=cours_plans_mapping,
                           generate_form=generate_form
                           )

@programme_bp.route('/<int:programme_id>/grille/generate', methods=['POST'])
@login_required
@ensure_profile_completed
def generate_programme_grille(programme_id):
    programme = Programme.query.get_or_404(programme_id)
    if current_user.role not in ('admin', 'coordo'):
        return jsonify({'error': 'Accès restreint (admin/coordo)'}), 403
    if current_user.role != 'admin' and programme not in current_user.programmes:
        return jsonify({'error': 'Accès refusé'}), 403
    user = current_user
    if not user.openai_key:
        return jsonify({'error': 'Aucune clé OpenAI configurée dans votre profil.'}), 400
    if not user.credits or user.credits <= 0:
        return jsonify({'error': "Crédits insuffisants pour effectuer l\'appel."}), 400

    form = request.get_json(silent=True) or {}
    from ..tasks.generation_grille import generate_programme_grille_task
    task = generate_programme_grille_task.delay(programme_id, user.id, form)
    return jsonify({'task_id': task.id}), 202

@programme_bp.route('/<int:programme_id>/competences/import_pdf/start', methods=['POST'])
@login_required
@ensure_profile_completed
def import_competences_pdf_start(programme_id):
    """Démarre l'import simplifié de compétences depuis un PDF pour un programme.

    Reçoit un formulaire multipart avec 'file' et retourne {task_id}.
    """
    from ..tasks.ocr import simple_import_competences_pdf
    programme = Programme.query.get_or_404(programme_id)
    if current_user.role not in ('admin', 'coordo') and programme not in current_user.programmes:
        return jsonify({'error': 'Accès refusé'}), 403
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni.'}), 400
    f = request.files['file']
    if not f or not f.filename.lower().endswith('.pdf'):
        return jsonify({'error': 'Veuillez fournir un fichier PDF.'}), 400
    upload_dir = os.path.join(current_app.config.get('UPLOAD_FOLDER', 'uploads'))
    os.makedirs(upload_dir, exist_ok=True)
    safe_name = re.sub(r'[^A-Za-z0-9_.-]+', '_', f.filename)
    filename = f"prog{programme_id}_{int(time.time())}_{safe_name}"
    pdf_path = os.path.join(upload_dir, filename)
    f.save(pdf_path)
    task = simple_import_competences_pdf.delay(programme_id, pdf_path, current_user.id)
    return jsonify({'task_id': task.id}), 202

@programme_bp.route('/<int:programme_id>/competences/import/review')
@login_required
@ensure_profile_completed
def review_competences_import_task(programme_id):
    """Affiche la page de comparaison à partir du résultat de tâche (task_id)."""
    from ...celery_app import celery
    from celery.result import AsyncResult
    programme = Programme.query.get_or_404(programme_id)
    task_id = request.args.get('task_id')
    if not task_id:
        flash('Identifiant de tâche manquant.', 'danger')
        return redirect(url_for('programme.view_competences_programme', programme_id=programme.id))
    res = AsyncResult(task_id, app=celery)
    data = res.result or {}
    base_filename = (data.get('result') or {}).get('base_filename') or ''

    comparisons = []
    try:
        txt_output_dir = current_app.config.get('TXT_OUTPUT_DIR', 'txt_outputs')
        source_list = (data.get('result') or {}).get('competences') or []
        if base_filename:
            json_path = os.path.join(txt_output_dir, f"{base_filename}_competences.json")
            if os.path.exists(json_path):
                with open(json_path, 'r', encoding='utf-8') as f:
                    j = json.load(f)
                    source_list = j.get('competences') or source_list
        for comp in source_list:
            code = comp.get('Code') or comp.get('code')
            if not code:
                continue
            db_comp = Competence.query.filter_by(code=code, programme_id=programme.id).first()
            db_version = None
            if db_comp:
                db_elements = []
                for elem in db_comp.elements:
                    criteria_list = [c.criteria for c in elem.criteria] if elem.criteria else []
                    db_elements.append({'nom': elem.nom, 'criteres': criteria_list})
                db_version = {
                    'code': db_comp.code,
                    'nom': db_comp.nom,
                    'contexte': db_comp.contexte_de_realisation,
                    'criteres': db_comp.criteria_de_performance,
                    'elements': db_elements,
                }
            json_elements = []
            elems = comp.get('Éléments') or comp.get('elements') or []
            for elem in elems:
                if isinstance(elem, str):
                    json_elements.append({'nom': elem, 'criteres': None})
                elif isinstance(elem, dict):
                    json_elements.append({'nom': elem.get('element') or elem.get('nom'), 'criteres': elem.get('criteres')})
            json_version = {
                'code': comp.get('Code') or comp.get('code'),
                'nom': comp.get('Nom de la compétence') or comp.get('nom'),
                'contexte': comp.get('Contexte de réalisation') or comp.get('Contexte') or comp.get('contexte_de_realisation'),
                'criteres': comp.get("Critères de performance pour l’ensemble de la compétence") or comp.get('criteria_de_performance'),
                'elements': json_elements,
            }
            comparisons.append({'code': code, 'db': db_version, 'json': json_version})
    except Exception:
        current_app.logger.exception('Erreur construction comparatif import (task)')
        comparisons = []

    form = ReviewImportConfirmForm()
    form.programme_id.data = programme_id
    form.base_filename.data = base_filename
    form.import_structured.data = 'true'
    return render_template('programme/review_import.html', programme=programme, comparisons=comparisons, base_filename=base_filename, form=form)

@programme_bp.route('/<int:programme_id>/grille/apply', methods=['POST'])
@login_required
@ensure_profile_completed
def apply_programme_grille(programme_id):
    programme = Programme.query.get_or_404(programme_id)
    if current_user.role not in ('admin', 'coordo'):
        return jsonify({'error': 'Accès restreint (admin/coordo)'}), 403
    if current_user.role != 'admin' and programme not in current_user.programmes:
        return jsonify({'error': 'Accès refusé'}), 403

    payload = request.get_json(silent=True) or {}
    mode = (payload.get('mode') or 'append').lower()
    grid = payload.get('grid') or {}
    sessions = grid.get('sessions') or []
    fils = grid.get('fils_conducteurs') or []

    # Option overwrite: supprimer les associations Cours↔Programme (pas les cours)
    if mode == 'overwrite':
        try:
            # Détacher tous les cours de ce programme
            for assoc in list(programme.cours_assocs):
                db.session.delete(assoc)
            db.session.flush()
        except Exception as e:
            db.session.rollback()
            return jsonify({'error': f"Erreur lors de l\'effacement de la grille: {e}"}), 500

    import random, string, re
    from ..models import Cours, CoursProgramme as CP, CoursPrealable, CoursCorequis

    def gen_unique_code():
        """Génère un code de cours unique préfixé par le code du programme.

        Si le programme possède un code ministériel (ex: 243.A0), on extrait le
        préfixe numérique (ex: 243) et on produit des codes du type '243-ABC'.
        À défaut, on utilise l'identifiant du programme comme préfixe.
        """
        # Déterminer le préfixe de programme
        prefix = None
        try:
            lp = getattr(programme, 'liste_programme_ministeriel', None)
            if lp and getattr(lp, 'code', None):
                m = re.match(r"^(\d{2,4})", lp.code.strip())
                if m:
                    prefix = m.group(1)
            if not prefix:
                prefix = str(programme.id)
        except Exception:
            prefix = str(programme.id)

        # Génère un code court aléatoire et vérifie l'unicité
        for _ in range(12):
            code = f"{prefix}-{''.join(random.choices(string.ascii_uppercase, k=3))}"
            if not Cours.query.filter_by(code=code).first():
                return code
        # Fallback numériques
        for _ in range(12):
            code = f"{prefix}-{random.randint(100,999)}"
            if not Cours.query.filter_by(code=code).first():
                return code
        # Dernier recours
        return f"{prefix}-{random.randint(1000,9999)}"

    created = []
    # Mapping temporaire: nom du cours -> (id, session)
    name_to_course = {}
    try:
        for sess_obj in sessions:
            sess_num = int(sess_obj.get('session') or 0)
            for course in (sess_obj.get('courses') or []):
                nom = (course.get('nom') or 'Cours généré').strip()
                ht = int(course.get('heures_theorie') or 0)
                hl = int(course.get('heures_laboratoire') or 0)
                hm = int(course.get('heures_travail_maison') or 0)
                unites = float(course.get('nombre_unites') or 0.0)
                if unites == 0.0:
                    try:
                        unites = round((ht + hl + hm), 2)
                    except Exception:
                        unites = float(ht + hl + hm) if all(isinstance(x, int) for x in (ht, hl, hm)) else 1.0

                c = Cours(code=gen_unique_code(), nom=nom,
                          nombre_unites=unites,
                          heures_theorie=ht,
                          heures_laboratoire=hl,
                          heures_travail_maison=hm)
                db.session.add(c)
                db.session.flush()
                # Associer au programme avec session
                assoc = CP()
                assoc.cours_id = c.id
                assoc.programme_id = programme.id
                assoc.session = sess_num
                db.session.add(assoc)
                created.append({'id': c.id, 'code': c.code, 'nom': c.nom, 'session': sess_num})
                # Conserver le mapping sur le nom pour les liens (préalables/corequis)
                # Si doublon de nom, on garde le premier rencontré pour la session correspondante
                name_key = c.nom.strip()
                name_to_course.setdefault(name_key, (c.id, sess_num))

        # Deuxième passe: appliquer préalables et co‑requis
        # Éviter doublons exacts en mémoire avant commit
        prealables_to_add = []  # tuples (cours_id, prealable_id)
        corequis_to_add = []    # tuples (cours_id, corequis_id)

        for sess_obj in sessions:
            sess_num = int(sess_obj.get('session') or 0)
            for course in (sess_obj.get('courses') or []):
                nom = (course.get('nom') or '').strip()
                if not nom or nom not in name_to_course:
                    continue
                cours_id, cours_session = name_to_course[nom]
                # Sanity: on attend que cours_session == sess_num
                # Gérer préalables (max 2, sessions antérieures seulement)
                raw_p = course.get('prealables') or course.get('prerequis') or []
                if isinstance(raw_p, str):
                    raw_p = [raw_p]
                added = 0
                seen = set()
                for p in raw_p:
                    pname = str(p).strip()
                    if not pname or pname.lower() in seen:
                        continue
                    seen.add(pname.lower())
                    ref = name_to_course.get(pname)
                    if not ref:
                        continue
                    ref_id, ref_session = ref
                    if ref_id == cours_id:
                        continue  # pas d'auto-référence
                    if ref_session >= cours_session:
                        continue  # préalables doivent être dans une session antérieure
                    prealables_to_add.append((cours_id, ref_id))
                    added += 1
                    if added >= 2:
                        break  # ne jamais dépasser 2 préalables

                # Gérer co‑requis (même session uniquement)
                raw_c = course.get('corequis') or []
                if isinstance(raw_c, str):
                    raw_c = [raw_c]
                seen_c = set()
                for co in raw_c:
                    cname = str(co).strip()
                    if not cname or cname.lower() in seen_c:
                        continue
                    seen_c.add(cname.lower())
                    ref = name_to_course.get(cname)
                    if not ref:
                        continue
                    ref_id, ref_session = ref
                    if ref_id == cours_id:
                        continue  # pas d'auto-référence
                    if ref_session != cours_session:
                        continue  # co‑requis dans la même session uniquement
                    corequis_to_add.append((cours_id, ref_id))

        # Insérer en base (en filtrant doublons)
        seen_pairs_p = set()
        for (cid, pid) in prealables_to_add:
            key = (cid, pid)
            if key in seen_pairs_p:
                continue
            seen_pairs_p.add(key)
            db.session.add(CoursPrealable(cours_id=cid, cours_prealable_id=pid, note_necessaire=None))

        seen_pairs_c = set()
        for (cid, coid) in corequis_to_add:
            key = (cid, coid)
            if key in seen_pairs_c:
                continue
            seen_pairs_c.add(key)
            db.session.add(CoursCorequis(cours_id=cid, cours_corequis_id=coid))

        # Appliquer les fils conducteurs (facultatif)
        try:
            from ..models import FilConducteur
            # Index existants par (description lower) pour éviter doublons
            existing_fils = { (f.description or '').strip().lower(): f for f in FilConducteur.query.filter_by(programme_id=programme.id).all() }
            for f in (fils if isinstance(fils, list) else []):
                try:
                    desc = str(f.get('description') or '').strip()
                    if not desc:
                        continue
                    color = str(f.get('couleur') or '').strip() or None
                    course_names = f.get('cours') or []
                    if isinstance(course_names, str):
                        course_names = [course_names]
                    # Trouver ou créer le fil
                    key = desc.lower()
                    fil = existing_fils.get(key)
                    if not fil:
                        fil = FilConducteur(programme_id=programme.id, description=desc, couleur=color)
                        db.session.add(fil)
                        db.session.flush()
                        existing_fils[key] = fil
                    elif color and (fil.couleur or '').strip() != color:
                        fil.couleur = color
                    # Associer les cours par nom
                    for cname in course_names:
                        try:
                            n = str(cname).strip()
                        except Exception:
                            continue
                        ref = name_to_course.get(n)
                        if not ref:
                            continue
                        cid, _ = ref
                        cobj = db.session.get(Cours, cid)
                        if cobj:
                            cobj.fil_conducteur_id = fil.id
                except Exception:
                    continue
        except Exception as e:
            # On n'interrompt pas l'application de la grille si la section fils échoue
            current_app.logger.warning(f"Application partielle: erreur sur fils conducteurs: {e}")

        db.session.commit()
        return jsonify({'ok': True, 'created': created}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f"Erreur lors de l\'application de la grille: {e}"}), 500


@programme_bp.route('/<int:programme_id>/grille/review/<task_id>')
@login_required
@ensure_profile_completed
def review_programme_grille(programme_id, task_id):
    """Page de validation/aperçu d'une grille générée (via task_id)."""
    programme = Programme.query.get_or_404(programme_id)
    if current_user.role not in ('admin', 'coordo'):
        flash('Accès restreint (admin/coordo).', 'danger')
        return redirect(url_for('main.index'))
    if current_user.role != 'admin' and programme not in current_user.programmes:
        flash('Accès refusé.', 'danger')
        return redirect(url_for('main.index'))

    # La page charge les détails via /tasks/status/<task_id> côté client et affiche un aperçu + bouton d'application
    return render_template('programme/review_grille_generation.html', programme=programme, task_id=task_id)

@programme_bp.route('/<int:programme_id>/grille/pdf')
@login_required
def export_programme_grille_pdf(programme_id):
    """Exporte la grille de cours du programme en PDF."""
    programme = Programme.query.get_or_404(programme_id)
    if programme not in current_user.programmes and current_user.role != 'admin':
        flash("Vous n'avez pas accès à ce programme.", 'danger')
        return redirect(url_for('main.index'))

    pdf_bytes = generate_programme_grille_pdf(programme)
    filename = f"grille_{programme.nom}.pdf"
    return send_file(
        io.BytesIO(pdf_bytes),
        mimetype='application/pdf',
        as_attachment=True,
        download_name=filename
    )

@programme_bp.route('/competence/<int:competence_id>/edit', methods=['GET', 'POST'])
@roles_required('admin', 'coordo')
@ensure_profile_completed
def edit_competence(competence_id):
    competence = Competence.query.get_or_404(competence_id)
    programmes = Programme.query.all()
    form = CompetenceForm()
    form.programmes.choices = [(p.id, p.nom) for p in programmes]

    if form.validate_on_submit():
        try:
            # MAJ many-to-many
            competence.programmes = Programme.query.filter(
                Programme.id.in_(form.programmes.data)
            ).all()
            competence.code = form.code.data
            competence.nom  = form.nom.data
            competence.criteria_de_performance   = form.criteria_de_performance.data
            competence.contexte_de_realisation   = form.contexte_de_realisation.data

            db.session.commit()
            flash('Compétence mise à jour avec succès !', 'success')

            # redirige vers le premier programme associé
            first_id = competence.programmes[0].id if competence.programmes else None
            return redirect(url_for('programme.view_programme', programme_id=first_id))

        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la mise à jour : {e}', 'danger')
            return redirect(url_for('programme.edit_competence', competence_id=competence_id))

    # --- GET ou validaton KO ---
    # pré-remplissage
    form.programmes.data             = [p.id for p in competence.programmes]
    form.code.data                   = competence.code or ''
    form.nom.data                    = competence.nom
    form.criteria_de_performance.data = competence.criteria_de_performance
    form.contexte_de_realisation.data = competence.contexte_de_realisation

    return render_template(
        'edit_competence.html',
        form=form,  # ou form=form
        competence=competence
    )


@programme_bp.route('/competence/<int:competence_id>/delete', methods=['POST'])
@roles_required('admin', 'coordo')
@ensure_profile_completed
def delete_competence(competence_id):
    # Formulaire de suppression
    delete_form = DeleteForm(prefix=f"competence-{competence_id}")

    if delete_form.validate_on_submit():
        competence = db.session.get(Competence, competence_id)
        if not competence:
            flash('Compétence non trouvée.', 'danger')
            return redirect(url_for('main.index'))

        programme_id = competence.programme_id
        try:
            db.session.delete(competence)
            db.session.commit()
            flash('Compétence supprimée avec succès!', 'success')
        except Exception as e:
            flash(f'Erreur lors de la suppression de la compétence : {e}', 'danger')

        return redirect(url_for('programme.view_programme', programme_id=programme_id))
    else:
        flash('Formulaire de suppression invalide.', 'danger')
        return redirect(url_for('main.index'))


@programme_bp.route('/competence/code/<string:competence_code>')
@role_required('admin')
@ensure_profile_completed
def view_competence_by_code(competence_code):
    # Récupérer la compétence par son code
    competence = Competence.query.filter_by(code=competence_code).first()

    if not competence:
        flash('Compétence non trouvée.', 'danger')
        return redirect(url_for('main.index'))

    # Rediriger vers la route existante avec competence_id
    return redirect(url_for('programme.view_competence', competence_id=competence.id))


@programme_bp.route('/competence/<int:competence_id>')
@login_required
@ensure_profile_completed
def view_competence(competence_id):
    # Récupération de la compétence + programme lié
    competence = db.session.get(Competence, competence_id)
    if not competence:
        flash('Compétence non trouvée.', 'danger')
        return redirect(url_for('main.index'))

    # Nettoyage HTML
    allowed_tags = [
        'ul', 'ol', 'li', 'strong', 'em', 'p', 'br', 'a',
        'h1', 'h2', 'h3', 'h4', 'h5', 'h6'
    ]
    allowed_attributes = {
        'a': ['href', 'title', 'target']
    }

    criteria_content = competence.criteria_de_performance or ""
    context_content = competence.contexte_de_realisation or ""

    # Assurer que ce sont des chaînes
    if not isinstance(criteria_content, str):
        criteria_content = str(criteria_content)
    if not isinstance(context_content, str):
        context_content = str(context_content)

    try:
        criteria_html = bleach.clean(
            criteria_content,
            tags=allowed_tags,
            attributes=allowed_attributes,
            strip=True
        )
        context_html = bleach.clean(
            context_content,
            tags=allowed_tags,
            attributes=allowed_attributes,
            strip=True
        )
    except Exception as e:
        flash(f'Erreur lors du nettoyage du contenu : {e}', 'danger')
        criteria_html = ""
        context_html = ""

    # Récupérer les éléments de compétence et leurs critères
    # Dans le modèle SQLAlchemy, un Competence possède un relationship 'elements'
    # et chaque ElementCompetence possède un relationship 'criteria'
    elements_comp = competence.elements  # liste de ElementCompetence

    # Organiser les critères par élément de compétence
    elements_dict = {}
    for element in elements_comp:
        elements_dict[element.id] = {
            'nom': element.nom,
            'criteres': [c.criteria for c in element.criteria]
        }

    # Instancier le formulaire de suppression
    delete_form = DeleteForm(prefix=f"competence-{competence.id}")

    return render_template(
        'view_competence.html',
        competence=competence,
        criteria_html=criteria_html,
        context_html=context_html,
        elements_competence=elements_dict,
        delete_form=delete_form
    )
